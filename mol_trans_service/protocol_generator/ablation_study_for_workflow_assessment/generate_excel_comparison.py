#!/usr/bin/env python3
"""
Generate an Excel file comparing cond_a vs cond_b workflow generation results.

Layout per sheet:
    Row 1 (merged): test_paper_1 (2 cols) | test_paper_2 (2 cols) | ...
    Row 2:          cond_a | cond_b       | cond_a | cond_b       | ...
    Row 3 (tall):   §6 Input Preparation content for each condition
    Row 4:          (separator)
    Row 5 (tall):   §7 Computational Workflow content for each condition

Usage:
    python generate_excel_comparison.py
    python generate_excel_comparison.py -m llama-3.1-70b gpt-4o
"""
import argparse
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

_CURRENT_DIR = Path(__file__).parent.resolve()
OUTPUTS_DIR = _CURRENT_DIR / "outputs"
TEST_QUERIES_PATH = _CURRENT_DIR / "test_queries.json"

MAX_CASE_INDEX = 15
FONT_NAME = "Times New Roman"
COND_A_LABEL = "MST-only"
COND_B_LABEL = "MST+Example"


def load_test_queries():
    with open(TEST_QUERIES_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    queries = data["test_queries"][:MAX_CASE_INDEX]
    return [(q["name"], q["label"]) for q in queries]


def find_reports_json(model_dir: Path, condition: str):
    run_dir = model_dir / condition / "run_1"
    if not run_dir.exists():
        return None
    report_files = sorted(run_dir.glob("*_reports.json"))
    return report_files[-1] if report_files else None


def load_reports(path):
    if path is None or not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {r["name"]: r.get("report_text", "") for r in data.get("results", [])}


def split_sections(report_text: str):
    """Split report into (section_6, section_7)."""
    if not report_text:
        return ("", "")

    s6_pattern = re.compile(r'^#{1,2}\s*6[\.\s]', re.MULTILINE)
    s7_pattern = re.compile(r'^#{1,2}\s*7[\.\s]', re.MULTILINE)

    s6_match = s6_pattern.search(report_text)
    s7_match = s7_pattern.search(report_text)

    if s6_match and s7_match and s6_match.start() < s7_match.start():
        section_6 = report_text[s6_match.start():s7_match.start()].strip()
        section_7 = report_text[s7_match.start():].strip()
    elif s6_match and not s7_match:
        section_6 = report_text[s6_match.start():].strip()
        section_7 = ""
    elif s7_match and not s6_match:
        section_6 = ""
        section_7 = report_text[s7_match.start():].strip()
    else:
        section_6 = ""
        section_7 = report_text.strip()

    return (section_6, section_7)


def add_model_sheet(wb: Workbook, model_name: str, model_dir: Path,
                    query_info: list):
    # Load reports
    cond_a_reports = load_reports(find_reports_json(model_dir, "cond_a"))
    cond_b_reports = load_reports(find_reports_json(model_dir, "cond_b"))

    if not cond_a_reports and not cond_b_reports:
        print(f"  ⚠️  {model_name}: no report data found, skipping")
        return

    ws = wb.create_sheet(title=model_name[:31])

    # Styles
    header_font = Font(name=FONT_NAME, bold=True, size=11)
    sub_header_font = Font(name=FONT_NAME, bold=True, size=10)
    row_label_font = Font(name=FONT_NAME, bold=True, size=11)
    body_font = Font(name=FONT_NAME, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    cond_a_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    cond_b_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")
    top_align = Alignment(vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    num_queries = len(query_info)

    # ── Row 1: Merged headers for each test_paper ──
    for q_idx, (name, _) in enumerate(query_info):
        col_a = 2 + q_idx * 2      # cond_a column
        col_b = col_a + 1           # cond_b column
        ws.merge_cells(
            start_row=1, start_column=col_a,
            end_row=1, end_column=col_b,
        )
        cell = ws.cell(row=1, column=col_a, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        # Border on the right cell of the merge
        ws.cell(row=1, column=col_b).border = thin_border

    # ── Row 2: cond_a / cond_b sub-headers ──
    for q_idx in range(num_queries):
        col_a = 2 + q_idx * 2
        col_b = col_a + 1

        cell_a = ws.cell(row=2, column=col_a, value=COND_A_LABEL)
        cell_a.font = sub_header_font
        cell_a.fill = cond_a_fill
        cell_a.alignment = center_align
        cell_a.border = thin_border

        cell_b = ws.cell(row=2, column=col_b, value=COND_B_LABEL)
        cell_b.font = sub_header_font
        cell_b.fill = cond_b_fill
        cell_b.alignment = center_align
        cell_b.border = thin_border

    # ── Row labels (column A) ──
    # Row 1-2: empty labels
    ws.cell(row=1, column=1, value="").border = thin_border
    ws.cell(row=2, column=1, value="").border = thin_border

    # Row 3: §6 Input Preparation
    label_s6 = ws.cell(row=3, column=1, value="§6 Input Preparation")
    label_s6.font = row_label_font
    label_s6.fill = header_fill
    label_s6.alignment = top_align
    label_s6.border = thin_border

    # Row 4: separator
    ws.cell(row=4, column=1, value="").border = thin_border

    # Row 5: §7 Computational Workflow
    label_s7 = ws.cell(row=5, column=1, value="§7 Computational Workflow")
    label_s7.font = row_label_font
    label_s7.fill = header_fill
    label_s7.alignment = top_align
    label_s7.border = thin_border

    # ── Fill content ──
    for q_idx, (name, _) in enumerate(query_info):
        col_a = 2 + q_idx * 2
        col_b = col_a + 1

        # Get section texts
        a_s6, a_s7 = split_sections(cond_a_reports.get(name, ""))
        b_s6, b_s7 = split_sections(cond_b_reports.get(name, ""))

        # Row 3: §6 Input Preparation
        cell = ws.cell(row=3, column=col_a, value=a_s6)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.fill = cond_a_fill
        cell.border = thin_border

        cell = ws.cell(row=3, column=col_b, value=b_s6)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.fill = cond_b_fill
        cell.border = thin_border

        # Row 4: separator
        ws.cell(row=4, column=col_a).border = thin_border
        ws.cell(row=4, column=col_b).border = thin_border

        # Row 5: §7 Computational Workflow
        cell = ws.cell(row=5, column=col_a, value=a_s7)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.fill = cond_a_fill
        cell.border = thin_border

        cell = ws.cell(row=5, column=col_b, value=b_s7)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.fill = cond_b_fill
        cell.border = thin_border

    # ── Sizing ──
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, 2 + num_queries * 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 45

    ws.row_dimensions[3].height = 350   # §6
    ws.row_dimensions[4].height = 15    # separator
    ws.row_dimensions[5].height = 500   # §7

    print(f"  ✅ {model_name}: {num_queries} queries")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel comparing cond_a vs cond_b workflow results"
    )
    parser.add_argument(
        "-m", "--models", nargs="+", default=None,
        help="Model names to include (default: all models)",
    )
    parser.add_argument(
        "-o", "--output", type=str, default=None,
        help="Output Excel path",
    )
    args = parser.parse_args()

    if not OUTPUTS_DIR.exists():
        print(f"❌ outputs/ not found at {OUTPUTS_DIR}")
        sys.exit(1)

    query_info = load_test_queries()

    if args.models:
        model_dirs = []
        for name in args.models:
            d = OUTPUTS_DIR / name
            if d.is_dir():
                model_dirs.append(d)
            else:
                print(f"⚠️  Not found: {d}")
        model_dirs.sort(key=lambda d: d.name)
    else:
        model_dirs = sorted([
            d for d in OUTPUTS_DIR.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        ])

    if not model_dirs:
        print("❌ No model directories found.")
        sys.exit(1)

    print(f"Processing {len(model_dirs)} model(s): {[d.name for d in model_dirs]}")

    wb = Workbook()
    wb.remove(wb.active)

    for model_dir in model_dirs:
        add_model_sheet(wb, model_dir.name, model_dir, query_info)

    output_path = Path(args.output) if args.output else OUTPUTS_DIR / "comparison_cond_a_vs_cond_b.xlsx"
    wb.save(output_path)
    print(f"\n📊 Excel saved to: {output_path}")


if __name__ == "__main__":
    main()
