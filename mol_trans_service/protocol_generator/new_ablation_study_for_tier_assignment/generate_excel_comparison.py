#!/usr/bin/env python3
"""
Generate an Excel file comparing ex1 vs ex3 results across all models.

Scans outputs/ for model directories, loads result JSON and reports JSON
from ex1/run_1/ and ex3/run_1/, and writes one sheet per model.

Layout per sheet (test_papers as columns):
    Row 1: expert_tier
    Row 2: ex1_predicted
    Row 3: ex1_report
    Row 4: ex3_predicted
    Row 5: ex3_report

Usage:
    # All models
    python generate_excel_comparison.py

    # Specific models only
    python generate_excel_comparison.py -m llama-3.1-70b gpt-4o
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
except ImportError:
    print("❌ openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

_CURRENT_DIR = Path(__file__).parent.resolve()
OUTPUTS_DIR = _CURRENT_DIR / "outputs"


def find_result_and_report(model_dir: Path, condition: str):
    """Find the result JSON and reports JSON for a condition/run_1."""
    run_dir = model_dir / condition / "run_1"
    if not run_dir.exists():
        return None, None

    # Result JSON (not _reports)
    result_files = [
        f for f in sorted(run_dir.glob("batch_test_*.json"))
        if "_reports" not in f.stem
    ]
    # Reports JSON
    report_files = sorted(run_dir.glob("*_reports.json"))

    result_path = result_files[-1] if result_files else None
    report_path = report_files[-1] if report_files else None
    return result_path, report_path


def load_json(path: Path):
    if path is None or not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_data(model_dir: Path):
    """Build structured data for one model.

    Returns:
        (query_names, expert_tiers, ex1_predicted, ex1_reports, ex3_predicted, ex3_reports)
        All dicts keyed by query name.
    """
    # Load ex1
    ex1_result_path, ex1_report_path = find_result_and_report(model_dir, "ex1")
    ex1_result_data = load_json(ex1_result_path)
    ex1_report_data = load_json(ex1_report_path)

    # Load ex3
    ex3_result_path, ex3_report_path = find_result_and_report(model_dir, "ex3")
    ex3_result_data = load_json(ex3_result_path)
    ex3_report_data = load_json(ex3_report_path)

    # Build lookup dicts
    query_names = []
    expert_tiers = {}
    ex1_predicted = {}
    ex3_predicted = {}

    for r in ex1_result_data.get("results", []):
        name = r["name"]
        if name not in expert_tiers:
            query_names.append(name)
        expert_tiers[name] = r.get("expert_tier", "")
        ex1_predicted[name] = r.get("predicted_tier", "")

    for r in ex3_result_data.get("results", []):
        name = r["name"]
        if name not in expert_tiers:
            query_names.append(name)
        expert_tiers[name] = r.get("expert_tier", "")
        ex3_predicted[name] = r.get("predicted_tier", "")

    # Report texts
    ex1_reports = {}
    for r in ex1_report_data.get("results", []):
        ex1_reports[r["name"]] = r.get("report_text", "")

    ex3_reports = {}
    for r in ex3_report_data.get("results", []):
        ex3_reports[r["name"]] = r.get("report_text", "")

    return query_names, expert_tiers, ex1_predicted, ex1_reports, ex3_predicted, ex3_reports


def add_model_sheet(wb: Workbook, model_name: str, model_dir: Path):
    """Add a sheet for one model."""
    query_names, expert_tiers, ex1_pred, ex1_rpts, ex3_pred, ex3_rpts = build_data(model_dir)

    if not query_names:
        print(f"  ⚠️  {model_name}: no data found, skipping")
        return

    # Sanitize sheet name (max 31 chars, no special chars)
    sheet_name = model_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Styles
    header_font = Font(name="Times New Roman", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row_label_font = Font(name="Times New Roman", bold=True, size=10)
    body_font = Font(name="Times New Roman", size=11)
    correct_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrong_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    staged_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    top_align = Alignment(vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Row 1: Header row with query names
    ws.cell(row=1, column=1, value="").font = header_font
    for col_idx, name in enumerate(query_names, start=2):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = top_align

    # Row labels
    row_labels = ["expert_tier", "ex1_predicted", "ex1_report", "ex3_predicted", "ex3_report"]
    row_data = [expert_tiers, ex1_pred, ex1_rpts, ex3_pred, ex3_rpts]

    for row_offset, (label, data_dict) in enumerate(zip(row_labels, row_data)):
        row_num = row_offset + 2

        # Row label cell
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = row_label_font
        label_cell.fill = header_fill
        label_cell.border = thin_border
        label_cell.alignment = top_align

        for col_idx, name in enumerate(query_names, start=2):
            value = data_dict.get(name, "")
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = body_font
            cell.border = thin_border

            if "report" in label:
                cell.alignment = wrap_align
            else:
                cell.alignment = top_align

            # Color predicted tier cells based on correctness
            if label in ("ex1_predicted", "ex3_predicted"):
                expert = expert_tiers.get(name, "")
                if value and expert:
                    if str(value).startswith("STAGED"):
                        cell.fill = staged_fill
                    elif value == expert:
                        cell.fill = correct_fill
                    else:
                        cell.fill = wrong_fill

    # Set column widths
    ws.column_dimensions["A"].width = 16
    for col_idx in range(2, len(query_names) + 2):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40

    # Set row heights for report rows
    ws.row_dimensions[4].height = 300  # ex1_report
    ws.row_dimensions[6].height = 300  # ex3_report

    print(f"  ✅ {model_name}: {len(query_names)} queries")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel comparing ex1 vs ex3 results"
    )
    parser.add_argument(
        "-m", "--models", nargs="+", default=None,
        help="Model names to include (default: all models in outputs/)",
    )
    parser.add_argument(
        "-o", "--output", type=str, default=None,
        help="Output Excel path (default: outputs/comparison_ex1_vs_ex3.xlsx)",
    )
    args = parser.parse_args()

    if not OUTPUTS_DIR.exists():
        print(f"❌ outputs/ directory not found at {OUTPUTS_DIR}")
        sys.exit(1)

    # Determine which model directories to process
    if args.models:
        model_dirs = []
        for name in args.models:
            d = OUTPUTS_DIR / name
            if d.is_dir():
                model_dirs.append(d)
            else:
                print(f"⚠️  Model directory not found: {d}")
        model_dirs.sort(key=lambda d: d.name)
    else:
        model_dirs = sorted([
            d for d in OUTPUTS_DIR.iterdir()
            if d.is_dir() and not d.name.startswith(".")
        ])

    if not model_dirs:
        print("❌ No model directories found.")
        sys.exit(1)

    print(f"Processing {len(model_dirs)} model(s): {[d.name for d in model_dirs]}")

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for model_dir in model_dirs:
        add_model_sheet(wb, model_dir.name, model_dir)

    output_path = Path(args.output) if args.output else OUTPUTS_DIR / "comparison_ex1_vs_ex3.xlsx"
    wb.save(output_path)
    print(f"\n📊 Excel saved to: {output_path}")


if __name__ == "__main__":
    main()
